#!/usr/bin/env python3
"""Read one PDF or XLSX and emit JSON Lines without modifying the source.

This worker is intentionally independent from DuckDB so it can run with the
bundled document runtime. ``build_silver.py`` consumes its records and writes
compact Parquet datasets.
"""

from __future__ import annotations

import argparse
import importlib.metadata
import json
import platform
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any


WORKER_VERSION = "1.0.0"


def clean_text(value: Any) -> str:
    rendered = ("" if value is None else str(value)).replace("\x00", "")
    return rendered.encode("utf-8", errors="replace").decode("utf-8")


def emit(record: dict[str, Any]) -> None:
    print(json.dumps(record, ensure_ascii=False, separators=(",", ":")))


def describe_environment() -> None:
    package_versions: dict[str, str | None] = {}
    for package in ("openpyxl", "pypdf", "pypdfium2"):
        try:
            package_versions[package] = importlib.metadata.version(package)
        except importlib.metadata.PackageNotFoundError:
            package_versions[package] = None

    tesseract = shutil.which("tesseract")
    tesseract_version: str | None = None
    tesseract_languages: list[str] = []
    if tesseract:
        version_process = subprocess.run(
            [tesseract, "--version"],
            capture_output=True,
            text=True,
            timeout=15,
            check=False,
        )
        if version_process.returncode == 0 and version_process.stdout:
            tesseract_version = version_process.stdout.splitlines()[0].strip()
        language_process = subprocess.run(
            [tesseract, "--list-langs"],
            capture_output=True,
            text=True,
            timeout=15,
            check=False,
        )
        if language_process.returncode == 0:
            tesseract_languages = sorted(
                line.strip()
                for line in language_process.stdout.splitlines()[1:]
                if line.strip()
            )

    emit(
        {
            "record_type": "environment",
            "worker_version": WORKER_VERSION,
            "python_version": platform.python_version(),
            "python_implementation": platform.python_implementation(),
            "platform": platform.platform(),
            "package_versions": package_versions,
            "tesseract_version": tesseract_version,
            "tesseract_languages": tesseract_languages,
        }
    )


def scalar_fields(value: Any) -> dict[str, Any]:
    if value is None:
        return {
            "value_kind": "null",
            "value_text": None,
            "numeric_value": None,
            "boolean_value": None,
            "date_value": None,
        }
    if isinstance(value, bool):
        return {
            "value_kind": "boolean",
            "value_text": "true" if value else "false",
            "numeric_value": None,
            "boolean_value": value,
            "date_value": None,
        }
    if isinstance(value, (int, float, Decimal)):
        return {
            "value_kind": "number",
            "value_text": clean_text(value),
            "numeric_value": clean_text(value),
            "boolean_value": None,
            "date_value": None,
        }
    if isinstance(value, (datetime, date)):
        return {
            "value_kind": "date",
            "value_text": value.isoformat(),
            "numeric_value": None,
            "boolean_value": None,
            "date_value": value.isoformat(),
        }
    return {
        "value_kind": "text",
        "value_text": clean_text(value),
        "numeric_value": None,
        "boolean_value": None,
        "date_value": None,
    }


def ocr_pdf_pages(path: Path, page_count: int, languages: str) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    texts = [""] * page_count
    tesseract = shutil.which("tesseract")
    if not tesseract:
        return texts, ["tesseract executable is unavailable"]
    try:
        import pypdfium2
    except ImportError:
        return texts, ["pypdfium2 is unavailable"]

    with tempfile.TemporaryDirectory(prefix="silver_pdf_ocr_") as directory:
        document = pypdfium2.PdfDocument(str(path))
        try:
            for index in range(page_count):
                image_path = Path(directory) / f"page_{index + 1:04d}.png"
                page = document[index]
                try:
                    bitmap = page.render(scale=2.0)
                    image = bitmap.to_pil()
                    image.save(image_path, format="PNG")
                    process = subprocess.run(
                        [
                            tesseract,
                            str(image_path),
                            "stdout",
                            "-l",
                            languages,
                            "--psm",
                            "6",
                        ],
                        capture_output=True,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        timeout=90,
                        check=False,
                    )
                    if process.returncode:
                        errors.append(
                            f"page {index + 1}: tesseract exit {process.returncode}: "
                            f"{process.stderr.strip()[:300]}"
                        )
                    else:
                        texts[index] = clean_text(process.stdout).strip()
                except Exception as error:  # keep other pages extractable
                    errors.append(f"page {index + 1}: {type(error).__name__}: {error}")
                finally:
                    page.close()
        finally:
            document.close()
    return texts, errors


def extract_pdf(path: Path, ocr_languages: str) -> None:
    import pypdf
    from pypdf import PdfReader

    reader = PdfReader(str(path), strict=False)
    encrypted = bool(reader.is_encrypted)
    if encrypted:
        decrypt_result = reader.decrypt("")
        if not decrypt_result:
            raise ValueError("PDF is encrypted and cannot be opened with an empty password")

    page_texts: list[str] = []
    page_methods: list[str] = []
    issues: list[dict[str, Any]] = []
    for index, page in enumerate(reader.pages, start=1):
        try:
            text = clean_text(page.extract_text() or "").strip()
        except Exception as error:
            text = ""
            issues.append(
                {
                    "record_type": "issue",
                    "severity": "error",
                    "unit_index": index,
                    "check_name": "pdf_page_text_extraction",
                    "message": f"{type(error).__name__}: {error}",
                }
            )
        page_texts.append(text)
        page_methods.append("pypdf")

    ocr_pages = 0
    text_total = sum(len(text) for text in page_texts)
    if page_texts and text_total < len(page_texts) * 40:
        ocr_texts, ocr_errors = ocr_pdf_pages(path, len(page_texts), ocr_languages)
        for index, ocr_text in enumerate(ocr_texts):
            if len(ocr_text) > len(page_texts[index]):
                page_texts[index] = ocr_text
                page_methods[index] = "tesseract_ocr"
                ocr_pages += 1
        for message in ocr_errors:
            issues.append(
                {
                    "record_type": "issue",
                    "severity": "warning",
                    "unit_index": None,
                    "check_name": "pdf_ocr",
                    "message": message,
                }
            )

    metadata: dict[str, Any] = {
        "library_versions": {"pypdf": pypdf.__version__},
    }
    if reader.metadata:
        metadata.update(
            {clean_text(key): clean_text(value) for key, value in reader.metadata.items()}
        )
    emit(
        {
            "record_type": "document",
            "format": "pdf",
            "worker_version": WORKER_VERSION,
            "page_count": len(page_texts),
            "sheet_count": None,
            "encrypted": encrypted,
            "ocr_pages": ocr_pages,
            "properties_json": json.dumps(metadata, ensure_ascii=False, sort_keys=True),
        }
    )
    for index, (text, method) in enumerate(zip(page_texts, page_methods), start=1):
        emit(
            {
                "record_type": "text_unit",
                "unit_type": "pdf_page",
                "unit_index": index,
                "unit_name": f"page_{index}",
                "extraction_method": method,
                "text": text,
            }
        )
    for issue in issues:
        emit(issue)


def extract_xlsx(path: Path, max_cells: int) -> None:
    import openpyxl

    formula_book = openpyxl.load_workbook(
        path, read_only=False, data_only=False, keep_links=False
    )
    value_book = openpyxl.load_workbook(
        path, read_only=False, data_only=True, keep_links=False
    )
    try:
        properties = {
            "library_versions": {"openpyxl": openpyxl.__version__},
            "creator": clean_text(formula_book.properties.creator),
            "title": clean_text(formula_book.properties.title),
            "subject": clean_text(formula_book.properties.subject),
            "created": formula_book.properties.created.isoformat()
            if formula_book.properties.created
            else None,
            "modified": formula_book.properties.modified.isoformat()
            if formula_book.properties.modified
            else None,
            "defined_name_count": len(formula_book.defined_names),
            "defined_name_broken_count": sum(
                1
                for value in formula_book.defined_names.values()
                if "#REF!" in clean_text(getattr(value, "attr_text", ""))
                or "#N/A" in clean_text(getattr(value, "attr_text", ""))
            ),
            "external_link_count": len(getattr(formula_book, "_external_links", [])),
        }
        emit(
            {
                "record_type": "document",
                "format": "xlsx",
                "worker_version": WORKER_VERSION,
                "page_count": None,
                "sheet_count": len(formula_book.worksheets),
                "encrypted": False,
                "ocr_pages": 0,
                "properties_json": json.dumps(properties, ensure_ascii=False, sort_keys=True),
            }
        )

        emitted_cells = 0
        for sheet_index, formula_sheet in enumerate(formula_book.worksheets, start=1):
            value_sheet = value_book[formula_sheet.title]
            potential_cells = formula_sheet.max_row * formula_sheet.max_column
            if emitted_cells + potential_cells > max_cells:
                emit(
                    {
                        "record_type": "issue",
                        "severity": "error",
                        "unit_index": sheet_index,
                        "check_name": "xlsx_cell_limit",
                        "message": (
                            f"Workbook exceeds configured cell scan limit {max_cells}; "
                            f"stopped before sheet {formula_sheet.title!r}."
                        ),
                    }
                )
                break

            formula_count = 0
            external_formula_count = 0
            nonempty_count = 0
            comment_count = 0
            actual_min_row: int | None = None
            actual_max_row: int | None = None
            actual_min_column: int | None = None
            actual_max_column: int | None = None
            merged_by_coordinate: dict[str, tuple[str, str]] = {}
            for merged in formula_sheet.merged_cells.ranges:
                anchor = formula_sheet.cell(merged.min_row, merged.min_col).coordinate
                for row_index in range(merged.min_row, merged.max_row + 1):
                    for column_index in range(merged.min_col, merged.max_col + 1):
                        coordinate = formula_sheet.cell(row_index, column_index).coordinate
                        merged_by_coordinate[coordinate] = (str(merged), anchor)
            for row in formula_sheet.iter_rows():
                for cell in row:
                    formula = clean_text(cell.value) if cell.data_type == "f" else None
                    cached = value_sheet[cell.coordinate].value if formula else None
                    effective = cached if formula and cached is not None else (
                        None if formula else cell.value
                    )
                    if formula is None and effective is None:
                        continue
                    formula_count += int(formula is not None)
                    formula_is_external = bool(
                        formula and re.search(r"\[[^\]]+\]", formula)
                    )
                    external_formula_count += int(formula_is_external)
                    nonempty_count += 1
                    comment_count += int(cell.comment is not None)
                    actual_min_row = cell.row if actual_min_row is None else min(actual_min_row, cell.row)
                    actual_max_row = cell.row if actual_max_row is None else max(actual_max_row, cell.row)
                    actual_min_column = (
                        cell.column
                        if actual_min_column is None
                        else min(actual_min_column, cell.column)
                    )
                    actual_max_column = (
                        cell.column
                        if actual_max_column is None
                        else max(actual_max_column, cell.column)
                    )
                    emitted_cells += 1
                    merged_range, merge_anchor = merged_by_coordinate.get(
                        cell.coordinate, (None, None)
                    )
                    emit(
                        {
                            "record_type": "cell",
                            "sheet_index": sheet_index,
                            "sheet_name": formula_sheet.title,
                            "row_index": cell.row,
                            "column_index": cell.column,
                            "cell_reference": cell.coordinate,
                            "formula": formula,
                            "formula_is_external": formula_is_external,
                            "cached_formula_value_text": clean_text(cached)
                            if cached is not None
                            else None,
                            "data_type": cell.data_type,
                            "number_format": clean_text(cell.number_format),
                            "style_id": int(cell.style_id),
                            "is_date": bool(cell.is_date),
                            "is_hidden_row": bool(
                                formula_sheet.row_dimensions[cell.row].hidden
                            ),
                            "is_hidden_column": bool(
                                formula_sheet.column_dimensions[cell.column_letter].hidden
                            ),
                            "merged_range": merged_range,
                            "merge_anchor": merge_anchor,
                            **scalar_fields(effective),
                        }
                    )

            hidden_rows = sum(
                1 for dimension in formula_sheet.row_dimensions.values() if dimension.hidden
            )
            hidden_columns = sum(
                int(dimension.max or dimension.min) - int(dimension.min) + 1
                for dimension in formula_sheet.column_dimensions.values()
                if dimension.hidden
            )
            emit(
                {
                    "record_type": "sheet",
                    "sheet_index": sheet_index,
                    "sheet_name": formula_sheet.title,
                    "sheet_state": formula_sheet.sheet_state,
                    "max_row": formula_sheet.max_row,
                    "max_column": formula_sheet.max_column,
                    "declared_range": formula_sheet.calculate_dimension(),
                    "actual_min_row": actual_min_row,
                    "actual_max_row": actual_max_row,
                    "actual_min_column": actual_min_column,
                    "actual_max_column": actual_max_column,
                    "nonempty_cell_count": nonempty_count,
                    "formula_count": formula_count,
                    "external_formula_count": external_formula_count,
                    "comment_count": comment_count,
                    "hidden_row_count": hidden_rows,
                    "hidden_column_count": hidden_columns,
                    "merged_ranges_json": json.dumps(
                        [str(value) for value in formula_sheet.merged_cells.ranges],
                        ensure_ascii=False,
                    ),
                }
            )
    finally:
        formula_book.close()
        value_book.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", type=Path)
    parser.add_argument("--format", choices=("pdf", "xlsx"))
    parser.add_argument("--ocr-languages", default="kor+eng+jpn")
    parser.add_argument("--max-cells", type=int, default=2_000_000)
    parser.add_argument("--describe-environment", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        if args.describe_environment:
            describe_environment()
            return 0
        if args.path is None or args.format is None:
            raise ValueError("--path and --format are required for extraction")
        if args.format == "pdf":
            extract_pdf(args.path, args.ocr_languages)
        else:
            extract_xlsx(args.path, args.max_cells)
        return 0
    except Exception as error:
        print(f"{type(error).__name__}: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
